import openpyxl

class RowColOperator:
  """This class operate row and column in worksheet of Excel."""

  def __init__(self):
    self.__ws = ''
    self.__target_row = 0
    self.__target_col = 0

  @property
  def get_ws(self):
    return self.__ws

  @property
  def get_target_row(self):
    return self.__target_row

  @property
  def get_target_col(self):
    return self.__target_col
  
  @get_ws.setter
  def set_ws(self, ws):
    self.__ws = ws

  @get_target_row.setter
  def set_target_row(self, target_row):
    self.__target_row = target_row

  @get_target_col.setter
  def set_target_col(self, target_col):
    self.__target_col = target_col

  def check_my_value_row(self):
    """Check my value to determine able to get end row in target sheet. 
    This function is for internal use only.

    Args:
        Nothing.

    Returns:
        bool: If able to get end row, return True.
    """
    if ('' != self.__ws) and (0 != self.__target_col): return True
    else: return False

  def check_my_value_col(self):
    """Check my value to determine able to get end column in target sheet. 
    This function is for internal use only.

    Args:
        Nothing.

    Returns:
        bool: If able to get end col, return True.
    """
    if ('' != self.__ws) and (0 != self.__target_row): return True
    else: return False

  def get_end_row(self) -> int:
    """This function gets end row in target worksheet.
    If target column values are all None, return 1.

    Note:
        The bellow variable must be set value before this function call.
        :obj:`self.__ws`, :obj:`self.__target_col`

    Args:
        Nothing.

    Returns:
        int: End row of the specified column in target worksheet.
        If target column values are all None, return 1.

    Examples:
        In case end row of 3rd column is 15.

        >>> print(row_col_operator.get_end_row())
        15

        ::

          row_col_operator = RowColOperator()
          row_col_operator.set_ws = openpyxl.load_workbook('workbook_name')['worksheet_name']
          row_col_operator.set_target_col = 3
          print(row_col_operator.get_end_row())
        
        ::
    """
    if not self.check_my_value_row(): return
    for cur_row in reversed(range(self.__target_col, self.__ws.max_row + 1)):
      if self.__ws.cell(row=cur_row, column=self.__target_col).value:
        return cur_row
    return 1

  def get_end_col(self) -> int:
    """This function gets end column in target worksheet.
    If target row values are all None, return 1.

    Note:
        The bellow variable must be set value before this function call.
        :obj:`self.__ws`, :obj:`self.__target_row`

    Args:
        Nothing.

    Returns:
        int: End column of the specified column in target worksheet.
        If target row values are all None, return 1.

    Examples:
        In case end column of 16th column is 4.

        >>> print(row_col_operator.get_end_col())
        4

        ::

          row_col_operator = RowColOperator()
          row_col_operator.set_ws = openpyxl.load_workbook('workbook_name')['worksheet_name']
          row_col_operator.set_target_row = 16
          print(row_col_operator.get_end_col())        

        ::
    """
    if not self.check_my_value_col(): return
    for cur_col in reversed(range(1, self.__ws.max_column + 1)):
      if self.__ws.cell(row=self.__target_row, column=cur_col).value:
        return cur_col
    return 1
