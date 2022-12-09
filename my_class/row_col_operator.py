import openpyxl

class RowColOperator:
  def __init__(self):
    self.__wb = ''
    self.__ws = ''
    self.__target_row = 0
    self.__target_col = 0
  
  #----------------------------------------------------------------
  #setter
  #----------------------------------------------------------------
  def set_wb(self, wb):
    self.__wb = wb

  def set_ws(self, ws):
    self.__ws = ws

  def set_target_row(self, target_row):
    self.__target_row = target_row

  def set_target_col(self, target_col):
    self.__target_col = target_col

  #----------------------------------------------------------------
  #getter
  #----------------------------------------------------------------
  def get_wb(self):
    return self.__wb

  def get_ws(self):
    return self.__ws

  def get_target_row(self):
    return self.__target_row

  def get_target_col(self):
    return self.__target_col
  
  #----------------------------------------------------------------
  #function
  #----------------------------------------------------------------
  def get_end_row(self):
    """get end row

    Returns:
        int: end row 
    """
    for cur_row in reversed(range(self.__target_col, self.__ws.max_row + 1)):
      if self.__ws.cell(row=cur_row, column=self.__target_col).value:
        return cur_row

  def get_end_col(self):
    """get end column

    Returns:
        int: end column
    """
    for cur_col in reversed(range(self.__target_row, self.__ws.max_column + 1)):
      if self.__ws.cell(row=self.__target_row, column=cur_col).value:
        return cur_col
