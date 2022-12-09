import openpyxl
from my_class import row_col_operator

# wb = openpyxl.load_workbook('test.xlsx')
# sheet = wb['Sheet1']
# sheet.cell(row=1, column=1).value = 'testtest'
# print(sheet['A1'].value)
# wb.save('test.xlsx')
# wb.close()

def main():
  RowColOperator = row_col_operator.RowColOperator()
  RowColOperator.set_wb(openpyxl.load_workbook('tool_openpyxl.xlsx'))
  RowColOperator.set_ws(RowColOperator.get_wb()['Sheet1'])
  RowColOperator.set_target_col(1)
  i = RowColOperator.get_end_row()
  RowColOperator.set_target_row(1)
  i = RowColOperator.get_end_col()


if __name__ == "__main__":
  main()